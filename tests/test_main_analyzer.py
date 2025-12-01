import os
from datetime import datetime, timedelta
from unittest.mock import patch

import pandas as pd
import pytest

from option_auditor.main_analyzer import analyze_csv
from option_auditor import main_analyzer as aud
from option_auditor.parsers import TastytradeParser, TastytradeFillsParser


def make_tasty_df(rows):
    # Minimal columns used by _normalize_tasty
    return pd.DataFrame(rows, columns=[
        "Time", "Underlying Symbol", "Quantity", "Action", "Price",
        "Commissions and Fees", "Expiration Date", "Strike Price", "Option Type"
    ])


def write_csv(df: pd.DataFrame, path):
    df.to_csv(path, index=False)
    return path


def test_detect_broker_tasty():
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.5,
            "Expiration Date": "2025-02-21",
            "Strike Price": 150,
            "Option Type": "Call",
        }
    ])
    assert aud._detect_broker(df) == "tasty"


def test_normalize_tasty_sign_and_proceeds():
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.5,
            "Expiration Date": "2025-02-21",
            "Strike Price": 150,
            "Option Type": "Call",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 2.0,
            "Commissions and Fees": 0.5,
            "Expiration Date": "2025-02-21",
            "Strike Price": 150,
            "Option Type": "Call",
        },
    ])
    parser = TastytradeParser()
    n = parser.parse(df)
    # First leg buy should be +qty, proceeds negative (outflow)
    assert n.loc[0, "qty"] == 1
    assert n.loc[0, "proceeds"] < 0
    # Second leg sell should be -qty in our signed convention
    assert n.loc[1, "qty"] == -1
    # Contract id built
    assert ":C:" in n.loc[0, "contract_id"]


def test_grouping_and_metrics_round_trip(tmp_path):
    # Build a simple profitable round-trip tasty CSV
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.00,
            "Commissions and Fees": 0.10,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
        {
            "Time": "2025-01-03 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.50,
            "Commissions and Fees": 0.10,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
    ])
    csv_path = write_csv(df, tmp_path / "tasty.csv")
    res = analyze_csv(str(csv_path), broker="tasty", account_size_start=10000, out_dir=str(tmp_path))
    assert res is not None


def test_excel_report_has_expected_sheets(tmp_path):
    # Minimal closed trade to trigger report
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open",
         "Price": 1.0, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Sell to Close",
         "Price": 1.2, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "excel.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    # Ensure file exists and sheets are present
    xlsx_path = tmp_path / "report.xlsx"
    assert os.path.exists(xlsx_path)
    try:
        import openpyxl
    except Exception:
        pytest.skip("openpyxl not available in environment")
    wb = openpyxl.load_workbook(str(xlsx_path))
    sheets = set(wb.sheetnames)
    # Summary must exist; and the three tabs
    assert "Summary" in sheets
    assert "Symbols" in sheets
    assert "Strategies" in sheets
    assert "Open Positions" in sheets


def test_csv_output_sanitization(tmp_path):
    # Ensure symbols/fields starting with =,+,-,@ are escaped in output CSV to prevent formula injection
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
    ])
    csv_path = write_csv(df, tmp_path / "inj.csv")
    analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    trades = pd.read_csv(tmp_path / "trades.csv")
    assert trades.loc[0, "symbol"].startswith("'")
    assert trades.loc[0, "contract_id"].startswith("'")


def test_symbol_aggregation_closed_only(tmp_path):
    # Build CSV with two symbols: MSFT closed round-trip, AAPL left open (single leg)
    df = make_tasty_df([
        # MSFT closed
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.00,
            "Commissions and Fees": 0.00,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.50,
            "Commissions and Fees": 0.00,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
        # AAPL open (only entry leg)
        {
            "Time": "2025-01-03 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 2.00,
            "Commissions and Fees": 0.00,
            "Expiration Date": "2025-03-21",
            "Strike Price": 150,
            "Option Type": "Call",
        },
    ])
    csv_path = write_csv(df, tmp_path / "agg.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=None)
    symbols = res.get("symbols", [])
    # Should contain MSFT only (closed), not AAPL (open)
    syms = {r["symbol"] for r in symbols}
    assert "MSFT" in syms
    assert "AAPL" not in syms
    # MSFT trades count 1 and positive pnl
    msft_row = next(r for r in symbols if r["symbol"] == "MSFT")
    assert msft_row["trades"] == 1
    assert msft_row["pnl"] > 0


def test_open_positions_listed(tmp_path):
    # Build CSV with one closed MSFT trade and one open AAPL position
    df = make_tasty_df([
        # Closed MSFT
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open",
         "Price": 1.00, "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Sell to Close",
         "Price": 1.20, "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        # Open AAPL (only entry leg)
        {"Time": "2025-01-03 10:00", "Underlying Symbol": "AAPL", "Quantity": 2, "Action": "Sell to Open",
         "Price": 1.50, "Commissions and Fees": 0.00, "Expiration Date": "2025-03-21", "Strike Price": 150, "Option Type": "Call"},
    ])
    csv_path = write_csv(df, tmp_path / "open_list.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=None)
    open_pos = res.get("open_positions", [])
    # Should contain AAPL with qty_open non-zero
    assert any(r.get("symbol") == "AAPL" and abs(float(r.get("qty_open", 0))) > 0 for r in open_pos)


def test_strategy_grouping_vertical_spread_classic_tasty(tmp_path):
    # Simulate an IWM bull put spread opened and later closed
    # Contract-level: one win (short), one loss (long) -> ~50% win rate
    # Strategy-level: spread net PnL positive -> 100% win rate
    rows = [
        # Open: collect 1.00 on short 229P, pay 0.50 on long 224P
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Sell to Open", "Price": 1.00,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 229, "Option Type": "Put"},
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Buy to Open", "Price": 0.50,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 224, "Option Type": "Put"},
        # Close later: buy back short at 0.20, sell long at 0.10
        {"Time": "2025-01-05 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Buy to Close", "Price": 0.20,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 229, "Option Type": "Put"},
        {"Time": "2025-01-05 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Sell to Close", "Price": 0.10,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 224, "Option Type": "Put"},
    ]
    df = make_tasty_df(rows)
    csv_path = write_csv(df, tmp_path / "spread.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=None)

    # Strategy-level metrics
    sm = res.get("strategy_metrics", {})
    assert sm.get("num_trades") == 1
    assert abs(sm.get("win_rate") - 1.0) < 1e-9  # 100%
    # Net PnL: (100 - 20) + (-50 + 10) = +40
    assert abs(sm.get("total_pnl") - 40.0) < 1e-6


def test_date_range_filtering_in_analyzer(tmp_path):
    # Two closed trades on different calendar days; filter should include only one day
    df = make_tasty_df([
        # Trade A (Jan 1 -> Jan 2)
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Sell to Close", "Price": 1.20,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        # Trade B (Jan 10 -> Jan 11)
        {"Time": "2025-01-10 10:00", "Underlying Symbol": "B", "Quantity": 1, "Action": "Buy to Open", "Price": 2.00,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-03-21", "Strike Price": 20, "Option Type": "Put"},
        {"Time": "2025-01-11 10:00", "Underlying Symbol": "B", "Quantity": 1, "Action": "Sell to Close", "Price": 2.50,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-03-21", "Strike Price": 20, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "range.csv")
    # Filter to only the first trade's dates
    res = analyze_csv(str(csv_path), broker="tasty", start_date="2025-01-01", end_date="2025-01-03", out_dir=None)
    assert len(res["strategy_groups"]) == 1
    # Filter to second trade window
    res2 = analyze_csv(str(csv_path), broker="tasty", start_date="2025-01-10", end_date="2025-01-12", out_dir=None)
    assert len(res2["strategy_groups"]) == 1


def test_correlation_matrix_calculation(tmp_path):
    # Trades across two symbols with both positive and negative correlation
    df = make_tasty_df([
        # Day 1: SPY win, QQQ win (positive correlation)
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        {"Time": "2025-01-01 12:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Sell to Open", "Price": 2.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 300, "Option Type": "Put"},
        {"Time": "2025-01-01 12:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Buy to Close", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 300, "Option Type": "Put"},
        # Day 2: SPY win, QQQ loss (negative correlation)
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 401, "Option Type": "Put"},
        {"Time": "2025-01-02 12:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 401, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 301, "Option Type": "Put"},
        {"Time": "2025-01-02 12:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Buy to Close", "Price": 1.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 301, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "corr.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    
    # Check JSON output
    corr_matrix = res.get("correlation_matrix")
    assert corr_matrix is not None
    assert len(corr_matrix) == 2
    
    # Check Excel output
    xlsx_path = tmp_path / "report.xlsx"
    assert os.path.exists(xlsx_path)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path))
        assert "Correlation" in wb.sheetnames
    except ImportError:
        pytest.skip("openpyxl not installed")

def test_map_sanitization(tmp_path):
    # Ensure symbols/fields starting with =,+,-,@ are escaped in output CSV to prevent formula injection
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
    ])
    csv_path = write_csv(df, tmp_path / "inj.csv")
    analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    trades = pd.read_csv(tmp_path / "trades.csv")
    assert trades.loc[0, "symbol"].startswith("'")
    assert trades.loc[0, "contract_id"].startswith("'")

def test_parse_tasty_datetime_invalid():
    parser = TastytradeParser()
    assert parser._parse_tasty_datetime("invalid date") is None

def test_normalize_tasty_fills_fallback_parser(tmp_path):
    # This test covers the fallback parser in _normalize_tasty_fills
    df = pd.DataFrame([
        {"Time": "2025-01-01 10:00", "Description": "1 Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "FALLBACK", "Commissions": 0, "Fees": 0}
    ])
    parser = TastytradeFillsParser()
    norm_df = parser.parse(df)
    assert not norm_df.empty
    assert norm_df.iloc[0]["symbol"] == "FALLBACK"
    assert norm_df.iloc[0]["strike"] == 100
    assert norm_df.iloc[0]["right"] == "C"

def test_verdict_logic(tmp_path):
    # Amber: win rate < 50% but positive PnL
    df1 = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Sell to Close", "Price": 2.00, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        {"Time": "2025-01-03 10:00", "Underlying Symbol": "B", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 20, "Option Type": "Call"},
        {"Time": "2025-01-04 10:00", "Underlying Symbol": "B", "Quantity": 1, "Action": "Sell to Close", "Price": 0.50, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 20, "Option Type": "Call"},
        {"Time": "2025-01-05 10:00", "Underlying Symbol": "C", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 30, "Option Type": "Call"},
        {"Time": "2025-01-06 10:00", "Underlying Symbol": "C", "Quantity": 1, "Action": "Sell to Close", "Price": 0.50, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 30, "Option Type": "Call"},
    ])
    csv_path1 = write_csv(df1, tmp_path / "amber.csv")
    res1 = analyze_csv(str(csv_path1))
    assert res1['verdict'] == "Amber"

    # Red flag: PnL < 0
    df2 = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Sell to Close", "Price": 0.50, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
    ])
    csv_path2 = write_csv(df2, tmp_path / "red.csv")
    res2 = analyze_csv(str(csv_path2))
    assert res2['verdict'] == "Red flag"

def test_normalize_tasty_fills_fallback_parser_edge_cases(tmp_path):
    parser = TastytradeFillsParser()
    # Missing quantity
    df1 = pd.DataFrame([{"Time": "2025-01-01 10:00", "Description": "Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "TEST", "Commissions": 0, "Fees": 0}])
    norm_df1 = parser.parse(df1)
    assert norm_df1.empty

    # Invalid quantity
    df2 = pd.DataFrame([{"Time": "2025-01-01 10:00", "Description": "X Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "TEST", "Commissions": 0, "Fees": 0}])
    norm_df2 = parser.parse(df2)
    assert norm_df2.empty

    # Infer expiry year
    df3 = pd.DataFrame([{"Time": "2024-11-01 10:00", "Description": "1 Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "TEST", "Commissions": 0, "Fees": 0}])
    norm_df3 = parser.parse(df3)
    assert norm_df3.iloc[0]["expiry"].year == 2025

def test_build_strategies_same_day_different_strategies(tmp_path):
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 400, "Option Type": "Call"},
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Close", "Price": 1.1, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 400, "Option Type": "Call"},
        {"Time": "2025-01-01 14:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 405, "Option Type": "Call"},
        {"Time": "2025-01-01 14:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Close", "Price": 1.1, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 405, "Option Type": "Call"},
    ])
    csv_path = write_csv(df, tmp_path / "multi_strat.csv")
    res = analyze_csv(str(csv_path))
    assert len(res["strategy_groups"]) == 2

def test_sym_desc_not_string():
    assert aud._sym_desc(123) == ""

def test_build_strategies_empty_df():
    from option_auditor.strategy import build_strategies
    df = pd.DataFrame(columns=["datetime", "contract_id", "symbol", "expiry", "strike", "right", "qty", "proceeds", "fees"])
    strategies = build_strategies(df)
    assert len(strategies) == 0

def test_sym_desc_unknown_symbol():
    assert aud._sym_desc("UNKNOWN") == "Options on UNKNOWN"

def test_analyze_csv_no_out_dir(tmp_path):
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Sell to Close", "Price": 1.2, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "test.csv")
    res = analyze_csv(str(csv_path), out_dir=None)
    assert "trades.csv" not in os.listdir(tmp_path)

def test_normalize_tasty_zero_quantity():
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 0, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
    ])
    parser = TastytradeParser()
    norm_df = parser.parse(df)
    assert norm_df.iloc[0]["qty"] == 0

def test_buying_power_calculation(tmp_path):
    csv_path = write_csv(make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
    ]), tmp_path / "dummy.csv")
    res = analyze_csv(str(csv_path), net_liquidity_now=10000, buying_power_available_now=2500)
    assert res["buying_power_utilized_percent"] == 75.0

    # Test division by zero case
    res_zero = analyze_csv(str(csv_path), net_liquidity_now=0, buying_power_available_now=0)
    assert res_zero["buying_power_utilized_percent"] is None

def test_rolling_trade_detection(tmp_path):
    df = make_tasty_df([
        # Original position
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        # Roll: close original, open new
        {"Time": "2025-01-05 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        {"Time": "2025-01-05 10:01", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.2, "Commissions and Fees": 0, "Expiration Date": "2025-02-20", "Strike Price": 400, "Option Type": "Put"},
        # Close rolled position
        {"Time": "2025-02-10 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.2, "Commissions and Fees": 0, "Expiration Date": "2025-02-20", "Strike Price": 400, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "roll.csv")
    res = analyze_csv(str(csv_path))
    assert len(res["strategy_groups"]) == 1
    assert "Rolled" in res["strategy_groups"][0]["strategy"]
    assert res["strategy_groups"][0]["pnl"] == (100 - 50) + (120 - 20)

def test_classify_strategy_three_legs():
    from option_auditor.strategy import _classify_strategy
    from option_auditor.models import StrategyGroup, TradeGroup, Leg
    strat = StrategyGroup(id="s1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    for i in range(3):
        leg = TradeGroup(contract_id=f"c{i}", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=100+i, right="C")
        leg.add_leg(Leg(ts=datetime.now(), qty=1, price=1, fees=0, proceeds=-100))
        strat.add_leg_group(leg)
    assert _classify_strategy(strat) == "Multi‑leg"

def test_invalid_strike_and_expiry():
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": "invalid", "Option Type": "Call"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Sell to Close", "Price": 2.00, "Commissions and Fees": 0, "Expiration Date": "invalid", "Strike Price": 10, "Option Type": "Call"},
    ])
    parser = TastytradeParser()
    norm_df = parser.parse(df)
    assert pd.isna(norm_df.iloc[0]["strike"])
    assert pd.isna(norm_df.iloc[1]["expiry"])

def test_group_contracts_with_open_empty_df():
    from option_auditor.main_analyzer import _group_contracts_with_open
    df = pd.DataFrame(columns=["datetime", "contract_id", "symbol", "expiry", "strike", "right", "qty", "proceeds", "fees"])
    closed, open = _group_contracts_with_open(df)
    assert len(closed) == 0
    assert len(open) == 0

def test_no_closed_trades(tmp_path):
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00, "Commissions and Fees": 0.10, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "open_only.csv")
    res = analyze_csv(str(csv_path))
    assert len(res["strategy_groups"]) == 1
    assert len(res["open_positions"]) == 1

def test_correlation_with_single_symbol(tmp_path):
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        {"Time": "2025-01-01 12:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "single_symbol.csv")
    res = analyze_csv(str(csv_path))
    assert res.get("correlation_matrix") is None
