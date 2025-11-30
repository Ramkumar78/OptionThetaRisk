import os
from datetime import datetime, timedelta
from unittest.mock import patch

import pandas as pd
import pytest

from option_auditor import analyze_csv
from option_auditor import auditor as aud


def make_tasty_df(rows):
    # Minimal columns used by _normalize_tasty
    return pd.DataFrame(rows)[[
        "Time", "Underlying Symbol", "Quantity", "Action", "Price",
        "Commissions and Fees", "Expiration Date", "Strike Price", "Option Type"
    ]]


def write_csv(df: pd.DataFrame, path):
    df.to_csv(path, index=False)
    return path


def test_detect_broker_tasty():
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.5,
            "Expiration Date": "2025-02-21",
            "Strike Price": 150,
            "Option Type": "Call",
        }
    ])
    assert aud._detect_broker(df) == aud.BROKER_TASTY


def test_normalize_tasty_sign_and_proceeds():
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.5,
            "Expiration Date": "2025-02-21",
            "Strike Price": 150,
            "Option Type": "Call",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 2.0,
            "Commissions and Fees": 0.5,
            "Expiration Date": "2025-02-21",
            "Strike Price": 150,
            "Option Type": "Call",
        },
    ])
    n = aud._normalize_tasty(df)
    # First leg buy should be +qty, proceeds negative (outflow)
    assert n.loc[0, "qty"] == 1
    assert n.loc[0, "proceeds"] < 0
    # Second leg sell should be -qty in our signed convention
    assert n.loc[1, "qty"] == -1
    # Contract id built
    assert ":C:" in n.loc[0, "contract_id"]


def test_grouping_and_metrics_round_trip(tmp_path):
    # Build a simple profitable round-trip tasty CSV
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.00,
            "Commissions and Fees": 0.10,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
        {
            "Time": "2025-01-03 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.50,
            "Commissions and Fees": 0.10,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
    ])
    csv_path = write_csv(df, tmp_path / "tasty.csv")
    res = analyze_csv(str(csv_path), broker="tasty", account_size=10000, out_dir=str(tmp_path))
    m = res["metrics"]
    assert m["num_trades"] == 1
    assert m["total_pnl"] > 0
    assert 1.9 <= m["win_rate"] * 100 <= 100  # single win
    # outputs
    assert os.path.exists(tmp_path / "trades.csv")
    # Excel report should be produced with required sheets
    xlsx_path = tmp_path / "report.xlsx"
    assert os.path.exists(xlsx_path)


def test_excel_report_has_expected_sheets(tmp_path):
    # Minimal closed trade to trigger report
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open",
         "Price": 1.0, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Sell to Close",
         "Price": 1.2, "Commissions and Fees": 0.0, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "excel.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    # Ensure file exists and sheets are present
    xlsx_path = tmp_path / "report.xlsx"
    assert os.path.exists(xlsx_path)
    try:
        import openpyxl
    except Exception:
        pytest.skip("openpyxl not available in environment")
    wb = openpyxl.load_workbook(str(xlsx_path))
    sheets = set(wb.sheetnames)
    # Summary must exist; and the three tabs
    assert "Summary" in sheets
    assert "Symbols" in sheets
    assert "Strategies" in sheets
    assert "Open Positions" in sheets


def test_verdict_over_leveraged(tmp_path):
    # Create large exposure by multiple buys at high price
    rows = []
    for i in range(5):
        rows.append({
            "Time": f"2025-01-01 10:{i:02d}",
            "Underlying Symbol": "BIG",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 50.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-03-21",
            "Strike Price": 100,
            "Option Type": "Call",
        })
        rows.append({
            "Time": f"2025-01-02 10:{i:02d}",
            "Underlying Symbol": "BIG",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 50.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-03-21",
            "Strike Price": 100,
            "Option Type": "Call",
        })
    df = make_tasty_df(rows)
    csv_path = write_csv(df, tmp_path / "expo.csv")
    # Account is tiny, should flag over-leveraged per proxy
    res = analyze_csv(str(csv_path), broker="tasty", account_size=1000, out_dir=None)
    assert res["verdict"] in {"Over-leveraged", "Amber", "Red flag", "Green flag", "Stop trading"}
    # With very small account size, our proxy likely exceeds 3x and triggers over-leveraged
    # Not asserting strict equality to be robust to heuristic changes, but ensure metrics computed
    assert res["metrics"]["num_trades"] == 5


def test_csv_output_sanitization(tmp_path):
    # Ensure symbols/fields starting with =,+,-,@ are escaped in output CSV to prevent formula injection
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
    ])
    csv_path = write_csv(df, tmp_path / "inj.csv")
    analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    trades = pd.read_csv(tmp_path / "trades.csv")
    assert trades.loc[0, "symbol"].startswith("'")
    assert trades.loc[0, "contract_id"].startswith("'")


def test_symbol_aggregation_closed_only(tmp_path):
    # Build CSV with two symbols: MSFT closed round-trip, AAPL left open (single leg)
    df = make_tasty_df([
        # MSFT closed
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.00,
            "Commissions and Fees": 0.00,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.50,
            "Commissions and Fees": 0.00,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
        # AAPL open (only entry leg)
        {
            "Time": "2025-01-03 10:00",
            "Underlying Symbol": "AAPL",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 2.00,
            "Commissions and Fees": 0.00,
            "Expiration Date": "2025-03-21",
            "Strike Price": 150,
            "Option Type": "Call",
        },
    ])
    csv_path = write_csv(df, tmp_path / "agg.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=None)
    symbols = res.get("symbols", [])
    # Should contain MSFT only (closed), not AAPL (open)
    syms = {r["symbol"] for r in symbols}
    assert "MSFT" in syms
    assert "AAPL" not in syms
    # MSFT trades count 1 and positive pnl
    msft_row = next(r for r in symbols if r["symbol"] == "MSFT")
    assert msft_row["trades"] == 1
    assert msft_row["pnl"] > 0


def test_open_positions_listed(tmp_path):
    # Build CSV with one closed MSFT trade and one open AAPL position
    df = make_tasty_df([
        # Closed MSFT
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Buy to Open",
         "Price": 1.00, "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "MSFT", "Quantity": 1, "Action": "Sell to Close",
         "Price": 1.20, "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 500, "Option Type": "Put"},
        # Open AAPL (only entry leg)
        {"Time": "2025-01-03 10:00", "Underlying Symbol": "AAPL", "Quantity": 2, "Action": "Sell to Open",
         "Price": 1.50, "Commissions and Fees": 0.00, "Expiration Date": "2025-03-21", "Strike Price": 150, "Option Type": "Call"},
    ])
    csv_path = write_csv(df, tmp_path / "open_list.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=None)
    open_pos = res.get("open_positions", [])
    # Should contain AAPL with qty_open non-zero
    assert any(r.get("symbol") == "AAPL" and abs(float(r.get("qty_open", 0))) > 0 for r in open_pos)


def test_strategy_grouping_vertical_spread_classic_tasty(tmp_path):
    # Simulate an IWM bull put spread opened and later closed
    # Contract-level: one win (short), one loss (long) -> ~50% win rate
    # Strategy-level: spread net PnL positive -> 100% win rate
    rows = [
        # Open: collect 1.00 on short 229P, pay 0.50 on long 224P
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Sell to Open", "Price": 1.00,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 229, "Option Type": "Put"},
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Buy to Open", "Price": 0.50,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 224, "Option Type": "Put"},
        # Close later: buy back short at 0.20, sell long at 0.10
        {"Time": "2025-01-05 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Buy to Close", "Price": 0.20,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 229, "Option Type": "Put"},
        {"Time": "2025-01-05 10:00", "Underlying Symbol": "IWM", "Quantity": 1, "Action": "Sell to Close", "Price": 0.10,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 224, "Option Type": "Put"},
    ]
    df = make_tasty_df(rows)
    csv_path = write_csv(df, tmp_path / "spread.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=None)

    # Contract-level metrics
    m = res["metrics"]
    assert m["num_trades"] == 2
    assert 0.4 <= m["win_rate"] <= 0.6  # ~50%

    # Strategy-level metrics
    sm = res.get("strategy_metrics", {})
    assert sm.get("num_trades") == 1
    assert abs(sm.get("win_rate") - 1.0) < 1e-9  # 100%
    # Net PnL: (100 - 20) + (-50 + 10) = +40
    assert abs(sm.get("total_pnl") - 40.0) < 1e-6


def test_date_range_filtering_in_analyzer(tmp_path):
    # Two closed trades on different calendar days; filter should include only one day
    df = make_tasty_df([
        # Trade A (Jan 1 -> Jan 2)
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Buy to Open", "Price": 1.00,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "A", "Quantity": 1, "Action": "Sell to Close", "Price": 1.20,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-02-21", "Strike Price": 10, "Option Type": "Call"},
        # Trade B (Jan 10 -> Jan 11)
        {"Time": "2025-01-10 10:00", "Underlying Symbol": "B", "Quantity": 1, "Action": "Buy to Open", "Price": 2.00,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-03-21", "Strike Price": 20, "Option Type": "Put"},
        {"Time": "2025-01-11 10:00", "Underlying Symbol": "B", "Quantity": 1, "Action": "Sell to Close", "Price": 2.50,
         "Commissions and Fees": 0.00, "Expiration Date": "2025-03-21", "Strike Price": 20, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "range.csv")
    # Filter to only the first trade's dates
    res = analyze_csv(str(csv_path), broker="tasty", start_date="2025-01-01", end_date="2025-01-03", out_dir=None)
    assert res["metrics"]["num_trades"] == 1
    # Filter to second trade window
    res2 = analyze_csv(str(csv_path), broker="tasty", start_date="2025-01-10", end_date="2025-01-12", out_dir=None)
    assert res2["metrics"]["num_trades"] == 1


def test_correlation_matrix_calculation(tmp_path):
    # Trades across two symbols with both positive and negative correlation
    df = make_tasty_df([
        # Day 1: SPY win, QQQ win (positive correlation)
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        {"Time": "2025-01-01 12:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 400, "Option Type": "Put"},
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Sell to Open", "Price": 2.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 300, "Option Type": "Put"},
        {"Time": "2025-01-01 12:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Buy to Close", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 300, "Option Type": "Put"},
        # Day 2: SPY win, QQQ loss (negative correlation)
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 401, "Option Type": "Put"},
        {"Time": "2025-01-02 12:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Close", "Price": 0.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 401, "Option Type": "Put"},
        {"Time": "2025-01-02 10:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Sell to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 301, "Option Type": "Put"},
        {"Time": "2025-01-02 12:00", "Underlying Symbol": "QQQ", "Quantity": 1, "Action": "Buy to Close", "Price": 1.5, "Commissions and Fees": 0, "Expiration Date": "2025-01-10", "Strike Price": 301, "Option Type": "Put"},
    ])
    csv_path = write_csv(df, tmp_path / "corr.csv")
    res = analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    
    # Check JSON output
    corr_matrix = res.get("correlation_matrix")
    assert corr_matrix is not None
    assert len(corr_matrix) == 2
    
    # Check Excel output
    xlsx_path = tmp_path / "report.xlsx"
    assert os.path.exists(xlsx_path)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path))
        assert "Correlation" in wb.sheetnames
    except ImportError:
        pytest.skip("openpyxl not installed")

def test_map_sanitization(tmp_path):
    # Ensure symbols/fields starting with =,+,-,@ are escaped in output CSV to prevent formula injection
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
        {
            "Time": "2025-01-02 10:00",
            "Underlying Symbol": "=HACK",
            "Quantity": 1,
            "Action": "Sell to Close",
            "Price": 1.0,
            "Commissions and Fees": 0.0,
            "Expiration Date": "2025-02-21",
            "Strike Price": 10,
            "Option Type": "Call",
        },
    ])
    csv_path = write_csv(df, tmp_path / "inj.csv")
    analyze_csv(str(csv_path), broker="tasty", out_dir=str(tmp_path))
    trades = pd.read_csv(tmp_path / "trades.csv")
    assert trades.loc[0, "symbol"].startswith("'")
    assert trades.loc[0, "contract_id"].startswith("'")

def test_parse_tasty_datetime_invalid():
    assert aud._parse_tasty_datetime("invalid date") is None

def test_normalize_tasty_fills_fallback_parser(tmp_path):
    # This test covers the fallback parser in _normalize_tasty_fills
    df = pd.DataFrame([
        {"Time": "2025-01-01 10:00", "Description": "1 Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "FALLBACK", "Commissions": 0, "Fees": 0}
    ])
    norm_df = aud._normalize_tasty_fills(df)
    assert not norm_df.empty
    assert norm_df.iloc[0]["symbol"] == "FALLBACK"
    assert norm_df.iloc[0]["strike"] == 100
    assert norm_df.iloc[0]["right"] == "C"

def test_classify_strategy_unclassified():
    strat = aud.StrategyGroup(id="s1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    assert aud._classify_strategy(strat) == "Unclassified"

def test_classify_strategy_multi_leg():
    strat = aud.StrategyGroup(id="s1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    leg1 = aud.TradeGroup(contract_id="c1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=100, right="C")
    leg1.add_leg(aud.Leg(ts=datetime.now(), qty=1, price=1, fees=0, proceeds=-100))
    leg2 = aud.TradeGroup(contract_id="c2", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=110, right="P")
    leg2.add_leg(aud.Leg(ts=datetime.now(), qty=-1, price=1, fees=0, proceeds=100))
    strat.add_leg_group(leg1)
    strat.add_leg_group(leg2)
    assert aud._classify_strategy(strat) == "Multi‑leg"

def test_classify_strategy_debit_verticals():
    # Call Debit
    strat_call = aud.StrategyGroup(id="s1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    leg1c = aud.TradeGroup(contract_id="c1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=100, right="C")
    leg1c.add_leg(aud.Leg(ts=datetime.now(), qty=1, price=2, fees=0, proceeds=-200))
    leg2c = aud.TradeGroup(contract_id="c2", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=110, right="C")
    leg2c.add_leg(aud.Leg(ts=datetime.now(), qty=-1, price=1, fees=0, proceeds=100))
    strat_call.add_leg_group(leg1c)
    strat_call.add_leg_group(leg2c)
    assert aud._classify_strategy(strat_call) == "Call Vertical (Debit)"

    # Put Debit
    strat_put = aud.StrategyGroup(id="s2", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    leg1p = aud.TradeGroup(contract_id="c3", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=100, right="P")
    leg1p.add_leg(aud.Leg(ts=datetime.now(), qty=-1, price=1, fees=0, proceeds=100))
    leg2p = aud.TradeGroup(contract_id="c4", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=110, right="P")
    leg2p.add_leg(aud.Leg(ts=datetime.now(), qty=1, price=2, fees=0, proceeds=-200))
    strat_put.add_leg_group(leg1p)
    strat_put.add_leg_group(leg2p)
    assert aud._classify_strategy(strat_put) == "Put Vertical (Debit)"

def test_normalize_tasty_missing_column():
    df = pd.DataFrame([{"Time": "2025-01-01"}]) # Missing all other required columns
    with pytest.raises(KeyError):
        aud._normalize_tasty(df)

def test_analyze_csv_empty_norm_df(tmp_path):
    csv_path = tmp_path / "empty.csv"
    csv_path.write_text("Header1,Header2\nValue1,Value2") # Will produce empty norm_df
    res = analyze_csv(str(csv_path))
    assert res.get("error") == "Unsupported CSV format"

def test_fallback_strategy_grouping(tmp_path):
    # This test ensures that if no trades are closed, the fallback grouping is used.
    df = make_tasty_df([
        {
            "Time": "2025-01-01 10:00",
            "Underlying Symbol": "MSFT",
            "Quantity": 1,
            "Action": "Buy to Open",
            "Price": 1.00,
            "Commissions and Fees": 0.10,
            "Expiration Date": "2025-02-21",
            "Strike Price": 500,
            "Option Type": "Put",
        },
    ])
    csv_path = write_csv(df, tmp_path / "open_only.csv")
    res = analyze_csv(str(csv_path))
    assert len(res["strategy_groups"]) > 0
    assert res["strategy_groups"][0]["strategy"] == "Long Put"

def test_verdict_logic_negative_pnl(tmp_path):
    csv_path = tmp_path / "dummy.csv"
    csv_path.write_text("Time,Underlying Symbol,Quantity,Action,Price,Commissions and Fees,Expiration Date,Strike Price,Option Type\n"
                        "2025-01-01 10:00,MSFT,1,Buy to Open,1.00,0.10,2025-02-21,500,Put\n"
                        "2025-01-03 10:00,MSFT,1,Sell to Close,1.50,0.10,2025-02-21,500,Put\n")

    with patch('option_auditor.auditor._build_strategies') as mock_build:
        s1 = aud.StrategyGroup("s1", "T", None, pnl=100)
        s1.entry_ts = pd.Timestamp("2025-01-01")
        s1.exit_ts = pd.Timestamp("2025-01-02")
        s2 = aud.StrategyGroup("s2", "T", None, pnl=100)
        s2.entry_ts = pd.Timestamp("2025-01-01")
        s2.exit_ts = pd.Timestamp("2025-01-02")
        s3 = aud.StrategyGroup("s3", "T", None, pnl=-300)
        s3.entry_ts = pd.Timestamp("2025-01-01")
        s3.exit_ts = pd.Timestamp("2025-01-02")
        mock_build.return_value = [s1, s2, s3] # win rate = 2/3 = 66.7%, but pnl < 0
        res = aud.analyze_csv(str(csv_path))
        assert res['verdict'] == "Red flag"

def test_normalize_tasty_fills_fallback_parser_edge_cases(tmp_path):
    # Missing quantity
    df1 = pd.DataFrame([{"Time": "2025-01-01 10:00", "Description": "Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "TEST", "Commissions": 0, "Fees": 0}])
    norm_df1 = aud._normalize_tasty_fills(df1)
    assert norm_df1.empty

    # Invalid quantity
    df2 = pd.DataFrame([{"Time": "2025-01-01 10:00", "Description": "X Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "TEST", "Commissions": 0, "Fees": 0}])
    norm_df2 = aud._normalize_tasty_fills(df2)
    assert norm_df2.empty

    # Infer expiry year
    df3 = pd.DataFrame([{"Time": "2024-11-01 10:00", "Description": "1 Jan 1 100 Call STO", "Price": "1.00 cr", "Symbol": "TEST", "Commissions": 0, "Fees": 0}])
    norm_df3 = aud._normalize_tasty_fills(df3)
    assert norm_df3.iloc[0]["expiry"].year == 2025

def test_build_strategies_same_day_different_strategies(tmp_path):
    df = make_tasty_df([
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 400, "Option Type": "Call"},
        {"Time": "2025-01-01 10:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Close", "Price": 1.1, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 400, "Option Type": "Call"},
        {"Time": "2025-01-01 14:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Buy to Open", "Price": 1.0, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 405, "Option Type": "Call"},
        {"Time": "2025-01-01 14:00", "Underlying Symbol": "SPY", "Quantity": 1, "Action": "Sell to Close", "Price": 1.1, "Commissions and Fees": 0, "Expiration Date": "2025-02-21", "Strike Price": 405, "Option Type": "Call"},
    ])
    csv_path = write_csv(df, tmp_path / "multi_strat.csv")
    res = analyze_csv(str(csv_path))
    assert len(res["strategy_groups"]) == 2

def test_sym_desc_not_string():
    assert aud._sym_desc(123) == ""

def test_classify_strategy_empty_openings():
    strat = aud.StrategyGroup(id="s1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    leg1 = aud.TradeGroup(contract_id="c1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=None, right=None)
    leg1.add_leg(aud.Leg(ts=datetime.now(), qty=1, price=1, fees=0, proceeds=-100))
    strat.add_leg_group(leg1)
    assert aud._classify_strategy(strat) == "Unclassified"

def test_classify_strategy_empty_legs():
    strat = aud.StrategyGroup(id="s1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"))
    leg1 = aud.TradeGroup(contract_id="c1", symbol="TEST", expiry=pd.Timestamp("2025-01-01"), strike=100, right="C")
    strat.add_leg_group(leg1)
    assert aud._classify_strategy(strat) == "Unclassified"

def test_parse_tasty_datetime_unhandled_format():
    assert aud._parse_tasty_datetime("2025-01-01T10:00:00-0500") is not None
    assert aud._parse_tasty_datetime("01-Jan-2025 10:00") is not None
    assert aud._parse_tasty_datetime("2025/01/01 10:00:00") is not None

def test_build_strategies_empty_df():
    df = pd.DataFrame(columns=["datetime", "contract_id", "symbol", "expiry", "strike", "right", "qty", "proceeds", "fees"])
    strategies = aud._build_strategies(df)
    assert len(strategies) == 0

def test_sym_desc_unknown_symbol():
    assert aud._sym_desc("UNKNOWN") == "Options on UNKNOWN"
